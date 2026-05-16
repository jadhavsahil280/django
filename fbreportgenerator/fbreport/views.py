from django.http import HttpResponse
from django.shortcuts import render
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl import load_workbook
from io import BytesIO

# Create your views here.
def index(request):


    if request.method == "POST":


        new_data = []

        file = request.FILES.get('file')

        df = pd.read_excel(file)

        rows = df.to_dict(orient='records')

        total_results = total_cpr = total_amount = total_facebookgst = total_agencycharges = 0


        for row in rows:
            adname = row.get('Ad name')
            addelivery = row.get("Ad delivery")

            results = float(row.get("Results")) if pd.notna(row.get("Results")) else 0
            total_results += results

            cpr = float(row.get("Cost per results")) if pd.notna(row.get("Cost per results")) else 0

            amountspent = float(row.get("Amount spent (INR)"))
            total_amount += amountspent

            facebookgst = amountspent*1.18
            total_facebookgst+=facebookgst

            agencycharges = (facebookgst*0.15)*1.18
            total_agencycharges += agencycharges

            # Store new row data
            new_data.append({   
                "Ad Name": adname,
                "Ad Delivery": addelivery,
                "Results": results,
                "Cost Per Result": cpr,
                "Amount Spent": amountspent,
                "Facebook GST": round(facebookgst, 2),
                "Agency Charges": round(agencycharges, 2),
            })


        total_cpr = total_amount / total_results

        new_data.append({   
                "Ad Name": "Total",
                "Ad Delivery": "",
                "Results": total_results,
                "Cost Per Result": total_cpr,
                "Amount Spent": total_amount,
                "Facebook GST": total_facebookgst,
                "Agency Charges": total_agencycharges,
            })
        
        new_data.append({   
                "Ad Name": "Total Amount Spend",
                "Ad Delivery": total_facebookgst + total_agencycharges
            })

        new_df = pd.DataFrame(new_data)

        # Create in-memory file
        output = BytesIO()


        # Save dataframe to excel
        with pd.ExcelWriter(output, engine='openpyxl') as writer:

            new_df.to_excel(writer, index=False, sheet_name='Report')

            workbook = writer.book
            worksheet = writer.sheets['Report']

            # Yellow background
            yellow_fill = PatternFill(
                start_color='FFFF00',
                end_color='FFFF00',
                fill_type='solid'
            )

            # Green background
            green_fill = PatternFill(
                start_color='00FF00',
                end_color='00FF00',
                fill_type='solid'
            )

            # Center alignment
            center_align = Alignment(
                horizontal='center',
                vertical='center'
            )

            # Bold font
            bold_font = Font(bold=True)

            # Apply style to header row
            for cell in worksheet[1]:
                cell.fill = yellow_fill
                cell.alignment = center_align
                cell.font = bold_font

            # Optional column auto width
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 5

            # Total row styling
            second_last_row = worksheet.max_row - 1

            for cell in worksheet[second_last_row]:
                cell.fill = yellow_fill
                cell.font = bold_font

            last_row = worksheet.max_row
            for cell in worksheet[last_row][:2]:
                cell.fill = green_fill
                cell.font = bold_font
                cell.alignment = center_align

            #round figure value
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, float) and not cell.value.is_integer():
                        cell.number_format = '0.00'

        output.seek(0)

        # Browser download dialog
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = 'attachment; filename="AD_Expense_Report.xlsx"'

        return response



    return render(request, 'fbreport/index.html')
